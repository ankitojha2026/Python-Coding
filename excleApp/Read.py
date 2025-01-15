from openpyxl import load_workbook
try:
    # Load the existing workbook
        wb = load_workbook("employee_data.xlsx")
        sheet = wb.active

        # Iterate through rows and print data
        print("Reading data:")
        for row in sheet.iter_rows(values_only=True):
            print(row[0])
except ModuleNotFoundError :
        print("error")
except PermissionError:
       print("permission denied")

