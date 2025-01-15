
import tkinter as tk
from tkinter import messagebox, filedialog
from openpyxl import load_workbook, Workbook

# Global variables to hold data from the sheets
file_first_data = []
file_second_data = []
win_width=1200
win_height=550


# Function to open file dialog and get file name for first sheet
def get_first_file_name():
    file_path = filedialog.askopenfilename(title="Select First Excel File", filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        return file_path
    else:
        messagebox.showerror("Error", "Please select a valid file")
        return None

# Function to read data from the first sheet
def read_sheet_one(file_name):
    try:
        wb = load_workbook(file_name)
        sheet = wb.active
        for row in sheet.iter_rows(values_only=True):
            file_first_data.append(row)
    except Exception as e:
        messagebox.showerror("Error", f"Error reading the first file: {e}")

# Function to open file dialog and get file name for second sheet
def get_second_file_name():
    file_path = filedialog.askopenfilename(title="Select Second Excel File", filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        return file_path
    else:
        messagebox.showerror("Error", "Please select a valid file")
        return None

# Function to read data from the second sheet
def read_sheet_two(file_name):
    try:
        wb = load_workbook(file_name)
        sheet = wb.active
        for row in sheet.iter_rows(values_only=True):
            file_second_data.append(row)
    except Exception as e:
        messagebox.showerror("Error", f"Error reading the second file: {e}")

# Function to write data to a new sheet
def add_into_new_Sheet(new_data, sheet):
    sheet.append(list(new_data))

# Function to save the new file with common data
def save_new_file():
    file_name = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    if file_name:
        wb = Workbook()
        sheet = wb.active
        for i in file_first_data:
            for j in file_second_data:
                if i[0] == j[0] and i[1] == j[1]:
                    add_into_new_Sheet(i, sheet)
        wb.save(file_name)
        messagebox.showinfo("Success", "Data filtered and saved successfully!")
    else:
        messagebox.showerror("Error", "Please choose a valid file name to save.")

# Function to handle the entire process of reading and comparing data
def process_files():
    # Get file names from the user
    first_file = get_first_file_name()
    if first_file:
        read_sheet_one(first_file)

    second_file = get_second_file_name()
    if second_file:
        read_sheet_two(second_file)

    if file_first_data and file_second_data:
        save_new_file()

# GUI setup using tkinter
def Excle_Tool():
    win = tk.Tk()
    win.title('Excle tool')
    win['bg'] = '#0AC1E9'
    win.resizable(False, False)
    screen_width = win.winfo_screenwidth()
    screen_height = win.winfo_screenheight()
    win.geometry(f'{win_width}x{win_height}+{(screen_width//2)-(win_width//2)}+{(screen_height//2)-(win_height//2)}')

    tk.Label(win,text='Excle tool', font=('Arial',30,'bold')).place(x=500,y=20)
    



    win.mainloop()
# Start the GUI application
Excle_Tool()
