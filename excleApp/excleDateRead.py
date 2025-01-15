# this linbraray is use for reading the data from the excle file
from openpyxl import load_workbook ,  Workbook
# this library is use for wriing into the new file
 #read the sheet one


 #we need two database for holding the data of the the both sheet 

file_first_data=[]
file_second_data=[]


#this function retun the name of the first file of the excel sheet
def get_first_file_name():
          file_name=input("First Excle Sheet Name: ")
          return file_name
     



def read_sheet_one():
    try:
        # Load the existing workbook
            # calling the get_first_file_name() return the name of the file
            file_name=get_first_file_name()
            wb = load_workbook(file_name+".xlsx")
            sheet = wb.active

            # Iterate through rows and print data
            for row in sheet.iter_rows(values_only=True):
                file_first_data.append(row)
    except ModuleNotFoundError :
            print("error")
    except PermissionError:
        print("permission denied")



def get_second_file_name():
      file_name=input("Second File Name: ")
      return file_name

# read sheet two

def read_sheet_two():
    
    try:
        # Load the existing workbook
            # calling the get_second_file_name() return the name of the file
            file_name=get_second_file_name()
            wb = load_workbook(file_name+".xlsx")
            sheet = wb.active

            # Iterate through rows and print data
            for row in sheet.iter_rows(values_only=True):
                file_second_data.append(row)
    except ModuleNotFoundError :
            print("error")
    except PermissionError:
        print("permission denied")


# adding into the new sheet
def add_into_new_Sheet(new_data,sheet):
      sheet.append(list(new_data))




#getnewfile name
def get_new_file_name():
      file=input("Enter New File Name For Save The New Excle :")
      return file


#checking the common rollnumber , gmail
def check_common_data():
        wb = Workbook()
        sheet = wb.active
        for i in file_first_data:
                for j in file_second_data:
                    if i[0]==j[0] and i[1]==j[1]:
                             add_into_new_Sheet(i,sheet)
        wb.save(get_new_file_name()+".xlsx") 
        print("Data filter success....")                      

# main function handle all things about the probgram
def main():
    #read sheet one
    read_sheet_one()
    # read sheet second 
    read_sheet_two()
    print("File is processing........ wait some sec....?")
    check_common_data()
main()